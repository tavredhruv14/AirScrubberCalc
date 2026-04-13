import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\Dhruv Tavre\.gemini\antigravity\Air Scrubber Altec working.xlsm', data_only=False)
ws = wb['Summary']

dump = {}
for row in ws.iter_rows(min_row=1, max_row=60):
    for cell in row:
        if cell.value is not None:
            dump[cell.coordinate] = cell.value

with open('summary_formulas.json', 'w') as f:
    json.dump(dump, f, indent=2)

wb_data = openpyxl.load_workbook(r'C:\Users\Dhruv Tavre\.gemini\antigravity\Air Scrubber Altec working.xlsm', data_only=True)
ws_data = wb_data['Summary']

dump_data = {}
for row in ws_data.iter_rows(min_row=1, max_row=60):
    for cell in row:
        if cell.value is not None:
            dump_data[cell.coordinate] = cell.value

with open('summary_data.json', 'w') as f:
    json.dump(dump_data, f, indent=2)

# Also dump GPM & Nozzle Suggested
ws2 = wb['GPM & Nozzle Suggested']
dump2 = {}
for row in ws2.iter_rows(min_row=1, max_row=40):
    for cell in row:
        if cell.value is not None:
            dump2[cell.coordinate] = cell.value
with open('gpm_nozzle_formulas.json', 'w') as f:
    json.dump(dump2, f, indent=2)
    
ws2_data = wb_data['GPM & Nozzle Suggested']
dump2_data = {}
for row in ws2_data.iter_rows(min_row=1, max_row=40):
    for cell in row:
        if cell.value is not None:
            dump2_data[cell.coordinate] = cell.value

with open('gpm_nozzle_data.json', 'w') as f:
    json.dump(dump2_data, f, indent=2)
