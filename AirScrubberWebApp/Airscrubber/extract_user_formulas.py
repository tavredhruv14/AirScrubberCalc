import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\Dhruv Tavre\.gemini\antigravity\Air Scrubber Altec working.xlsm', data_only=False)
ws = wb['Summary']

dump = {}
dump["E21"] = str(ws["E21"].value)
dump["G10"] = str(ws["G10"].value)

for row in range(1, 53):
    dump[f"D{row}"] = str(ws[f"D{row}"].value)
    dump[f"E{row}"] = str(ws[f"E{row}"].value)
    dump[f"G{row}"] = str(ws[f"G{row}"].value)

rows = {}
for r in range(11, 52):
    row_data = {}
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f"{col}{r}"]
        row_data[col] = str(cell.value)
    rows[r] = row_data

dump["table_formulas"] = rows

with open('user_formulas.json', 'w') as f:
    json.dump(dump, f, indent=2)
